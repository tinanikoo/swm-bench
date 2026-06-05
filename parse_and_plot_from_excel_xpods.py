#!/usr/bin/env python3
"""
Create xpods charts/plots directly from an input Excel file (no log parsing).

Expected input workbook:
  - sheet "Raw" and/or
  - sheet "Summary_means"
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import List, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.patches import Rectangle
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows


SUMMARY_REQUIRED_COLS = [
    "system",
    "pod_count",
    "qps",
    "PodScheduled_p99",
    "PodScheduled_p50",
    "PodScheduled_max",
    "PodScheduled_avg",
    "ContainerReady_p99",
    "ContainerReady_p50",
    "ContainerReady_max",
    "ContainerReady_avg",
    "ServiceLatency_99th",
    "ServiceLatency_p50",
    "ServiceLatency_max",
    "ServiceLatency_avg",
    "ResourceUsage_runningPods",
    "ResourceUsage_cpu_total_m",
    "ResourceUsage_mem_total_mi",
    "ResourceUsage_net_rx_bytes_total",
    "ResourceUsage_net_tx_bytes_total",
]


def _norm_sched_name(x: object) -> str:
    s = str(x).strip().lower()
    if "assured" in s:
        return "qos_assured"
    if "qos" in s:
        return "qos"
    if "default" in s or s == "def":
        return "def"
    return s


SYSTEM_ORDER = {"def": 0, "qos": 1, "qos_assured": 2}
SYSTEM_LABELS = {"def": "Default", "qos": "QoS", "qos_assured": "QoS-ASSURED"}
SYSTEM_COLORS = {"def": "blue", "qos": "green", "qos_assured": "orange"}


def _ordered_systems(summary: pd.DataFrame) -> list[str]:
    systems = [str(x) for x in summary["system"].dropna().unique().tolist()]
    systems = sorted(systems, key=lambda s: (SYSTEM_ORDER.get(s, 99), s))
    return systems


def _display_system(s: str) -> str:
    return SYSTEM_LABELS.get(s, s)


def _system_color(s: str) -> str:
    return SYSTEM_COLORS.get(s, "gray")


def print_run_count_table_from_raw(raw_df: pd.DataFrame) -> None:
    if raw_df.empty:
        print("[INFO] Run-count table unavailable: input workbook has no Raw sheet.")
        return

    tmp = raw_df.copy()
    if "scheduler" in tmp.columns and tmp["scheduler"].notna().any():
        tmp["sch"] = tmp["scheduler"].map(_norm_sched_name)
    elif "system" in tmp.columns:
        tmp["sch"] = tmp["system"].map(_norm_sched_name)
    else:
        print("[INFO] Run-count table unavailable: Raw sheet missing scheduler/system columns.")
        return

    counts = (
        tmp.groupby(["sch", "qps", "pod_count"], dropna=False)
        .size()
        .reset_index(name="run_count")
        .sort_values(["run_count", "sch", "qps", "pod_count"], ascending=[False, True, True, True], kind="stable")
        .reset_index(drop=True)
    )
    print("[INFO] Run-count table (from input Raw sheet):")
    print(counts.to_string(index=False))


def load_input_excel(path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    xl = pd.ExcelFile(path)
    raw_df = pd.DataFrame()
    summary_df = pd.DataFrame()

    if "Raw" in xl.sheet_names:
        raw_df = pd.read_excel(path, sheet_name="Raw")

    if "Summary_means" in xl.sheet_names:
        summary_df = pd.read_excel(path, sheet_name="Summary_means")

    if summary_df.empty and not raw_df.empty:
        metric_cols = [
            "PodScheduled_p99",
            "PodScheduled_p50",
            "PodScheduled_max",
            "PodScheduled_avg",
            "ContainerReady_p99",
            "ContainerReady_p50",
            "ContainerReady_max",
            "ContainerReady_avg",
            "ServiceLatency_99th",
            "ServiceLatency_p50",
            "ServiceLatency_max",
            "ServiceLatency_avg",
            "ResourceUsage_runningPods",
            "ResourceUsage_cpu_total_m",
            "ResourceUsage_mem_total_mi",
            "ResourceUsage_net_rx_bytes_total",
            "ResourceUsage_net_tx_bytes_total",
        ]
        metric_cols = [c for c in metric_cols if c in raw_df.columns]
        summary_df = (
            raw_df.groupby(["system", "pod_count", "qps"], as_index=False)[metric_cols]
            .mean(numeric_only=True)
            .round(2)
        )

    if summary_df.empty:
        raise SystemExit(
            "ERROR: input workbook must contain either sheet 'Summary_means' or sheet 'Raw' with parseable columns."
        )

    missing = [c for c in SUMMARY_REQUIRED_COLS if c not in summary_df.columns]
    hard_missing = [c for c in missing if not c.endswith("_p50") and not c.startswith("ResourceUsage_")]
    if hard_missing:
        raise SystemExit(f"ERROR: Summary data missing columns: {hard_missing}")
    for c in missing:
        summary_df[c] = np.nan

    for c in ["pod_count", "qps"]:
        summary_df[c] = pd.to_numeric(summary_df[c], errors="coerce")
    summary_df = summary_df.dropna(subset=["pod_count", "qps"]).copy()
    summary_df["pod_count"] = summary_df["pod_count"].astype(int)
    summary_df["qps"] = summary_df["qps"].astype(int)
    summary_df["system"] = summary_df["system"].astype(str).str.lower()

    summary_df = summary_df.sort_values(["pod_count", "qps", "system"], kind="stable").reset_index(drop=True)
    return raw_df, summary_df


def build_excel_with_summary_and_charts(raw_df: pd.DataFrame, summary: pd.DataFrame, out_xlsx: Path) -> None:
    raw_df = raw_df.where(pd.notna(raw_df), None)
    summary = summary.where(pd.notna(summary), None)

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw"
    ws_sum = wb.create_sheet("Summary_means")
    ws_ch = wb.create_sheet("Charts")

    if not raw_df.empty:
        for r in dataframe_to_rows(raw_df, index=False, header=True):
            ws_raw.append(r)
    else:
        ws_raw.append(["info"])
        ws_raw.append(["Raw sheet not provided in input workbook"])

    for r in dataframe_to_rows(summary, index=False, header=True):
        ws_sum.append(r)

    pod_counts = sorted(int(x) for x in summary["pod_count"].dropna().unique().tolist())
    qps_values = sorted(int(x) for x in summary["qps"].dropna().unique().tolist())
    systems = _ordered_systems(summary)

    metrics = {
        "PodScheduled": ("PodScheduled_avg", "PodScheduled_p50", "PodScheduled_p99", "PodScheduled_max"),
        "ContainerReady": ("ContainerReady_avg", "ContainerReady_p50", "ContainerReady_p99", "ContainerReady_max"),
        "ServiceLatency": ("ServiceLatency_avg", "ServiceLatency_p50", "ServiceLatency_99th", "ServiceLatency_max"),
    }
    resource_metrics = {
        "ResourceUsage_runningPods": "ResourceUsage_runningPods",
        "ResourceUsage_cpu_total_m": "ResourceUsage_cpu_total_m",
        "ResourceUsage_mem_total_mi": "ResourceUsage_mem_total_mi",
        "ResourceUsage_net_rx_bytes_total": "ResourceUsage_net_rx_bytes_total",
        "ResourceUsage_net_tx_bytes_total": "ResourceUsage_net_tx_bytes_total",
    }
    stats = [("Avg", 0), ("P50", 1), ("P99", 2), ("Max", 3)]

    def write_table(start_row: int, start_col: int, qps: int, metric_name: str, stat_label: str, stat_col: str):
        ws_ch.cell(row=start_row, column=start_col, value=f"QPS {qps} - {metric_name} ({stat_label})")
        ws_ch.cell(row=start_row + 1, column=start_col, value="pod_count")
        for j, sys in enumerate(systems):
            ws_ch.cell(row=start_row + 1, column=start_col + 1 + j, value=sys)

        for i, pods in enumerate(pod_counts):
            r = start_row + 2 + i
            ws_ch.cell(row=r, column=start_col, value=pods)

            for j, sys in enumerate(systems):
                v_s = summary.loc[
                    (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), stat_col
                ]
                s_val = v_s.iloc[0] if len(v_s) else None
                ws_ch.cell(row=r, column=start_col + 1 + j, value=float(s_val) if pd.notna(s_val) else None)

        header_row = start_row + 1
        last_row = start_row + 1 + len(pod_counts)
        return header_row, last_row

    def add_chart(
        table_header_row: int,
        table_left_col: int,
        table_last_row: int,
        title: str,
        anchor_cell: str,
        chart_kind: str,
    ):
        chart = BarChart() if chart_kind == "bar" else LineChart()
        chart.title = title
        chart.x_axis.title = "number of pods"
        chart.y_axis.title = "value (ms)"
        chart.legend.position = "r"
        if chart_kind == "bar":
            chart.type = "col"
            chart.style = 10

        cats = Reference(ws_ch, min_col=table_left_col, min_row=table_header_row + 1, max_row=table_last_row)
        data = Reference(
            ws_ch,
            min_col=table_left_col + 1,
            min_row=table_header_row,
            max_col=table_left_col + len(systems),
            max_row=table_last_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = False
        ws_ch.add_chart(chart, anchor_cell)

    row_cursor = 1
    left_col = 1
    for qps in qps_values:
        ws_ch.cell(row=row_cursor, column=left_col, value=f"QPS {qps} (QoS vs Default across number of pods)")
        row_cursor += 1
        for metric_name, cols in metrics.items():
            for stat_label, idx in stats:
                stat_col = cols[idx]
                top = row_cursor
                header_row, last_row = write_table(top, left_col, qps, metric_name, stat_label, stat_col)
                anchor = ws_ch.cell(row=top, column=left_col + 4).coordinate
                chart_kind = "bar" if stat_label == "P50" else "line"
                add_chart(header_row, left_col, last_row, f"QPS {qps} - {metric_name} ({stat_label})", anchor, chart_kind)
                row_cursor = last_row + 4
        for metric_name, stat_col in resource_metrics.items():
            top = row_cursor
            header_row, last_row = write_table(top, left_col, qps, metric_name, "Total", stat_col)
            anchor = ws_ch.cell(row=top, column=left_col + 4).coordinate
            add_chart(header_row, left_col, last_row, f"QPS {qps} - {metric_name} (Total)", anchor, "bar")
            row_cursor = last_row + 4
        row_cursor += 2

    for ws in [ws_raw, ws_sum, ws_ch]:
        max_col = min(ws.max_column, 22)
        for col_idx in range(1, max_col + 1):
            col_letter = chr(64 + col_idx)
            max_len = 0
            for row_idx in range(1, min(ws.max_row, 200) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(out_xlsx)


def _auto_log_scale(values: np.ndarray) -> bool:
    vals = values[np.isfinite(values) & (values > 0)]
    if len(vals) < 2:
        return False
    return (np.nanmax(vals) / np.nanmin(vals)) >= 50


def plot_boxrange_xpods(
    summary: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
    label_avg: bool = False,
    avg_label_decimals: int = 0,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    systems = _ordered_systems(summary)

    metric_map = {
        "PodScheduled": ("PodScheduled_avg", "PodScheduled_p50", "PodScheduled_p99", "PodScheduled_max"),
        "ContainerReady": ("ContainerReady_avg", "ContainerReady_p50", "ContainerReady_p99", "ContainerReady_max"),
        "ServiceLatency": ("ServiceLatency_avg", "ServiceLatency_p50", "ServiceLatency_99th", "ServiceLatency_max"),
    }
    box_width = 0.30
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, (avg_col, p50_col, p99_col, max_col) in metric_map.items():
            vals = []
            for sys in systems:
                sub = summary[
                    (summary["system"] == sys)
                    & (summary["qps"] == qps)
                    & (summary["pod_count"].isin(pod_counts))
                ]
                if not sub.empty:
                    vals.extend(sub[[avg_col, p50_col, p99_col, max_col]].to_numpy().ravel().tolist())
            all_vals = np.array([v for v in vals if v is not None], dtype=float)
            use_log = _auto_log_scale(all_vals)
            if use_log and not np.all(all_vals[np.isfinite(all_vals)] > 0):
                use_log = False

            fig, ax = plt.subplots(figsize=(7.8, 4.6))
            x = np.arange(len(pod_counts), dtype=float)
            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("Time (ms)")
            ax.set_title(f"QPS {qps} - {metric_name} (box=p99..max, dot=avg)")

            nsys = max(len(systems), 1)
            if nsys == 1:
                offsets = [0.0]
            else:
                span = 0.54
                step = span / (nsys - 1)
                offsets = [(-span / 2) + i * step for i in range(nsys)]

            for idx, sys in enumerate(systems):
                dx = offsets[idx]
                label = _display_system(sys)
                color = _system_color(sys)
                sub = summary[
                    (summary["system"] == sys)
                    & (summary["qps"] == qps)
                    & (summary["pod_count"].isin(pod_counts))
                ].copy()
                sub["pod_count"] = pd.Categorical(sub["pod_count"], categories=list(pod_counts), ordered=True)
                sub = sub.sort_values("pod_count")

                proxy_done = False
                for i, pods in enumerate(pod_counts):
                    row = sub[sub["pod_count"] == pods]
                    if row.empty:
                        continue
                    avg = row.iloc[0][avg_col]
                    p99 = row.iloc[0][p99_col]
                    mx = row.iloc[0][max_col]
                    if pd.isna(avg) and pd.isna(p99) and pd.isna(mx):
                        continue

                    xi = x[i] + dx
                    if pd.notna(p99) and pd.notna(mx) and float(mx) >= float(p99):
                        rect = Rectangle(
                            (xi - box_width / 2, float(p99)),
                            box_width,
                            float(mx) - float(p99),
                            fill=False,
                            linewidth=1.8,
                            edgecolor=color,
                            label=label if not proxy_done else None,
                        )
                        ax.add_patch(rect)
                        proxy_done = True

                    if pd.notna(avg):
                        y = float(avg)
                        ax.plot([xi], [y], marker="o", linestyle="None", color=color)
                        if label_avg:
                            fmt = f"{{:.{avg_label_decimals}f}}"
                            voff = 6 if idx % 2 == 0 else -10
                            ax.annotate(
                                fmt.format(y),
                                (xi, y),
                                textcoords="offset points",
                                xytext=(0, voff),
                                ha="center",
                                va="bottom" if voff > 0 else "top",
                                color=color,
                                fontsize=8,
                            )

            handles, labels = ax.get_legend_handles_labels()
            if handles:
                ax.legend(loc="best")
            ax.grid(True, which="both", axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            if use_log:
                ax.set_yscale("log")

            out_path = out_dir / f"boxrange_{metric_name}_qps{qps}.png"
            fig.tight_layout()
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_p50_bar_xpods(
    summary: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    systems = _ordered_systems(summary)
    metric_map = {
        "PodScheduled": "PodScheduled_p50",
        "ContainerReady": "ContainerReady_p50",
        "ServiceLatency": "ServiceLatency_p50",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, p50_col in metric_map.items():
            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), p50_col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))
            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("P50 (ms)")
            ax.set_title(f"QPS {qps} - {metric_name} (P50 bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            ax.legend(loc="best")

            out_path = out_dir / f"bar_p50_{metric_name}_qps{qps}.png"
            fig.tight_layout()
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_resource_bar_xpods(
    summary: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    systems = _ordered_systems(summary)
    metric_map = {
        "ResourceUsage_runningPods": "ResourceUsage_runningPods",
        "ResourceUsage_cpu_total_m": "ResourceUsage_cpu_total_m",
        "ResourceUsage_mem_total_mi": "ResourceUsage_mem_total_mi",
        "ResourceUsage_net_rx_bytes_total": "ResourceUsage_net_rx_bytes_total",
        "ResourceUsage_net_tx_bytes_total": "ResourceUsage_net_tx_bytes_total",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, col in metric_map.items():
            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))
            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel(metric_name)
            ax.set_title(f"QPS {qps} - {metric_name} (bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            ax.legend(loc="best")

            out_path = out_dir / f"bar_resource_{metric_name}_qps{qps}.png"
            fig.tight_layout()
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            created.append(out_path)
    return created


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Read normal.xlsx-like input and generate xpods charts and plots (no log parsing)."
    )
    parser.add_argument(
        "--in-excel",
        default="/home/samizadeh/swm-bench/bookinfo-codecoapp/log-result/normal.xlsx",
        help="Input workbook path (must include Summary_means or Raw).",
    )
    parser.add_argument(
        "--out-excel",
        default="normal_xpods_report.xlsx",
        help="Output Excel filename (placed next to input unless absolute path).",
    )
    parser.add_argument(
        "--plots-dir",
        default="plots_boxrange_from_excel_xpods",
        help="Output plots directory (placed next to input unless absolute path).",
    )
    parser.add_argument(
        "--p50-plots-dir",
        default="plots_p50_bar_from_excel_xpods",
        help="Output directory for additional P50 bar plots (placed next to input unless absolute path).",
    )
    parser.add_argument(
        "--resource-plots-dir",
        default="plots_resource_bar_from_excel_xpods",
        help="Output directory for ResourceUsage bar plots (placed next to input unless absolute path).",
    )
    parser.add_argument("--pods", default="", help="Optional comma-separated pod counts.")
    parser.add_argument("--qps", default="", help="Optional comma-separated qps values.")
    parser.add_argument("--label-avg", action="store_true", help="Write avg value label near each avg dot.")
    parser.add_argument("--avg-label-decimals", type=int, default=0, help="Decimals for avg labels.")
    args = parser.parse_args()

    in_excel = Path(args.in_excel).expanduser()
    if not in_excel.exists():
        raise SystemExit(f"ERROR: input excel not found: {in_excel}")

    base_dir = in_excel.parent

    def resolve_out(p: str) -> Path:
        pp = Path(p).expanduser()
        return pp if pp.is_absolute() else base_dir / pp

    raw_df, summary = load_input_excel(in_excel)
    print_run_count_table_from_raw(raw_df)

    out_excel = resolve_out(args.out_excel)
    build_excel_with_summary_and_charts(raw_df, summary, out_excel)
    print(f"[OK] Excel written: {out_excel}")

    if args.pods.strip():
        pod_counts = tuple(int(x.strip()) for x in args.pods.split(",") if x.strip())
    else:
        pod_counts = tuple(sorted(int(x) for x in summary["pod_count"].dropna().unique()))

    if args.qps.strip():
        qps_values = tuple(int(x.strip()) for x in args.qps.split(",") if x.strip())
    else:
        qps_values = tuple(sorted(int(x) for x in summary["qps"].dropna().unique()))

    plots_dir = resolve_out(args.plots_dir)
    created = plot_boxrange_xpods(
        summary=summary,
        out_dir=plots_dir,
        pod_counts=pod_counts,
        qps_values=qps_values,
        label_avg=args.label_avg,
        avg_label_decimals=args.avg_label_decimals,
    )
    print(f"[OK] {len(created)} plots written to: {plots_dir}")
    for p in created:
        print(f"  - {p}")

    p50_plots_dir = resolve_out(args.p50_plots_dir)
    created_p50 = plot_p50_bar_xpods(
        summary=summary,
        out_dir=p50_plots_dir,
        pod_counts=pod_counts,
        qps_values=qps_values,
    )
    print(f"[OK] {len(created_p50)} P50 bar plots written to: {p50_plots_dir}")
    for p in created_p50:
        print(f"  - {p}")

    resource_plots_dir = resolve_out(args.resource_plots_dir)
    created_resource = plot_resource_bar_xpods(
        summary=summary,
        out_dir=resource_plots_dir,
        pod_counts=pod_counts,
        qps_values=qps_values,
    )
    print(f"[OK] {len(created_resource)} ResourceUsage bar plots written to: {resource_plots_dir}")
    for p in created_resource:
        print(f"  - {p}")


if __name__ == "__main__":
    main()
