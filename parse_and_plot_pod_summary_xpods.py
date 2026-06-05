#!/usr/bin/env python3
"""
Parse bookinfo pod summary logs, build Excel report, and create PNG plots.

This variant uses:
  - X-axis: pod_count
  - Separation: qps and metric

It keeps the original plots and adds:
  - plots_boxrange_xpods2: summary box plots
  - plots_resource_bar_xpods2: alternative resource bar plots
"""

from __future__ import annotations

import argparse
import gc
import math
import re
from pathlib import Path
from typing import List, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.patches import Rectangle
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows

from parse_and_plot_pod_summary import (
    CONTAINER_READY_COUNTS_RE,
    METRIC_LINE_RE,
    parse_file,
    split_blocks,
    summarize_means,
    to_nullable_ints,
)


POD_SCHEDULED_COUNTS_RE = re.compile(
    r"PodScheduledTime[^\n]*\bexpectedPods=(?P<expected_pods>\d+)\b[^\n]*\bscheduledPods=(?P<scheduled_pods>\d+)\b"
)


def _norm_sched_name(x: object) -> str:
    s = str(x).strip().lower()
    if "assured" in s:
        return "qos_assured"
    if "qos" in s:
        return "qos"
    if "default" in s or s == "def":
        return "def"
    return s


SYSTEM_ORDER = {"def": 0, "qos": 1, "qos_assured": 2}
SYSTEM_LABELS = {"def": "Default", "qos": "QoS", "qos_assured": "QoS-ASSURED"}
SYSTEM_COLORS = {"def": "blue", "qos": "green", "qos_assured": "orange"}
NUMBER_OF_RUNS = 20


def _ordered_systems(summary: pd.DataFrame) -> list[str]:
    systems = [str(x) for x in summary["system"].dropna().unique().tolist()]
    systems = sorted(systems, key=lambda s: (SYSTEM_ORDER.get(s, 99), s))
    return systems


def _display_system(s: str) -> str:
    return SYSTEM_LABELS.get(s, s)


def _system_color(s: str) -> str:
    return SYSTEM_COLORS.get(s, "gray")


def print_run_count_table(raw_df: pd.DataFrame, label: str = "from input logs") -> None:
    if raw_df.empty:
        print("[INFO] Run-count table: no raw rows found.")
        return

    tmp = raw_df.copy()
    if "scheduler" in tmp.columns and tmp["scheduler"].notna().any():
        tmp["sch"] = tmp["scheduler"].map(_norm_sched_name)
    else:
        tmp["sch"] = tmp["system"].map(_norm_sched_name)

    counts = (
        tmp.groupby(["pod_count", "qps", "sch"], dropna=False)
        .size()
        .reset_index(name="run_count")
        .sort_values(["pod_count", "qps", "sch"], kind="stable")
        .reset_index(drop=True)
    )
    print(f"[INFO] Run-count table ({label}):")
    print(counts.to_string(index=False))
    filtered_pods = sorted(int(x) for x in tmp["pod_count"].dropna().unique())
    filtered_qps = sorted(int(x) for x in tmp["qps"].dropna().unique())
    print(f"[INFO] Filtered pods: {filtered_pods}")
    print(f"[INFO] Filtered qps: {filtered_qps}")


def _pod_scheduled_counts_for_file(path: Path) -> pd.DataFrame:
    text = path.read_text(encoding="utf-8", errors="replace")
    rows = []
    run_seq = 0

    for block in split_blocks(text):
        matches = list(METRIC_LINE_RE.finditer(block))
        if not matches:
            continue

        run_seq += 1
        pod_counts = POD_SCHEDULED_COUNTS_RE.search(block)
        if not pod_counts:
            continue

        rows.append(
            {
                "input_log": path.name,
                "run_id": int(matches[0].group("run_id")),
                "run_seq": run_seq,
                "qps": int(matches[0].group("qps")),
                "PodScheduled_expectedPods": int(pod_counts.group("expected_pods")),
                "PodScheduled_scheduledPods": int(pod_counts.group("scheduled_pods")),
            }
        )

    return pd.DataFrame(rows)


def _log_context_from_text(path: Path, text: str) -> dict[str, object]:
    qps_match = re.search(r"\bqps=(\d+)\b", text)
    pod_match = re.search(r"_(?:qos(?:-assured)?|default|def)-scheduler_(?P<pods>\d+)(?:_\d+)?\.log$", path.name)
    return {
        "input_log": path.name,
        "qps": int(qps_match.group(1)) if qps_match else None,
        "pod_count": int(pod_match.group("pods")) if pod_match else None,
    }


def filter_log_paths_by_readiness(log_paths: List[Path]) -> List[Path]:
    valid_paths: List[Path] = []
    invalid_rows = []

    for path in log_paths:
        text = path.read_text(encoding="utf-8", errors="replace")
        context = _log_context_from_text(path, text)
        reasons = []
        pod_counts = POD_SCHEDULED_COUNTS_RE.search(text)
        container_counts = CONTAINER_READY_COUNTS_RE.search(text)

        if not pod_counts:
            reasons.append("PST missing")
            context["expectedPods"] = None
            context["scheduledPods"] = None
        else:
            expected_pods = int(pod_counts.group("expected_pods"))
            scheduled_pods = int(pod_counts.group("scheduled_pods"))
            context["expectedPods"] = expected_pods
            context["scheduledPods"] = scheduled_pods
            if expected_pods != scheduled_pods:
                reasons.append("expectedPods != scheduledPods")

        if not container_counts:
            reasons.append("CRT missing")
            context["expectedContainers"] = None
            context["readyContainers"] = None
        else:
            expected_containers = int(container_counts.group("expected_containers"))
            ready_containers = int(container_counts.group("ready_containers"))
            context["expectedContainers"] = expected_containers
            context["readyContainers"] = ready_containers
            if expected_containers != ready_containers:
                reasons.append("expectedContainers != readyContainers")

        if reasons:
            context["reason"] = "; ".join(reasons)
            invalid_rows.append(context)
        else:
            valid_paths.append(path)

    print("[INFO] Filtering log phase (file readiness checks before parsing):")
    if invalid_rows:
        invalid_df = pd.DataFrame(invalid_rows)
        print(invalid_df.to_string(index=False))
        print(f"[INFO] Excluded {len(invalid_rows)} log files before parsing.")
    else:
        print("[INFO] No log files excluded before parsing.")

    return valid_paths


def add_pod_scheduled_counts(raw_df: pd.DataFrame, log_paths: List[Path]) -> pd.DataFrame:
    count_parts = [_pod_scheduled_counts_for_file(p) for p in log_paths]
    count_parts = [p for p in count_parts if not p.empty]
    if not count_parts:
        return raw_df

    counts = pd.concat(count_parts, ignore_index=True)
    return raw_df.merge(counts, on=["input_log", "run_id", "run_seq", "qps"], how="left")


def filter_readiness_runs(raw_df: pd.DataFrame) -> pd.DataFrame:
    checks = [
        (
            "PodScheduledTime missing",
            "expectedPods != scheduledPods",
            "PodScheduled_expectedPods",
            "PodScheduled_scheduledPods",
        ),
        (
            "ContainerReadyTime missing",
            "expectedContainers != readyContainers",
            "ContainerReady_expectedContainers",
            "ContainerReady_readyContainers",
        ),
    ]
    invalid_rows = []
    keep_mask = pd.Series(True, index=raw_df.index)

    for idx, row in raw_df.iterrows():
        reasons = []
        for missing_reason, mismatch_reason, expected_col, actual_col in checks:
            expected = pd.to_numeric(pd.Series([row.get(expected_col)]), errors="coerce").iloc[0]
            actual = pd.to_numeric(pd.Series([row.get(actual_col)]), errors="coerce").iloc[0]
            if pd.isna(expected) or pd.isna(actual):
                reasons.append(missing_reason)
            elif expected != actual:
                reasons.append(mismatch_reason)

        if reasons:
            keep_mask.loc[idx] = False
            invalid_rows.append(
                {
                    "input_log": row.get("input_log"),
                    "run_id": row.get("run_id"),
                    "qps": row.get("qps"),
                    "pod_count": row.get("pod_count"),
                    "expectedPods": row.get("PodScheduled_expectedPods"),
                    "scheduledPods": row.get("PodScheduled_scheduledPods"),
                    "expectedContainers": row.get("ContainerReady_expectedContainers"),
                    "readyContainers": row.get("ContainerReady_readyContainers"),
                    "reason": "; ".join(reasons),
                }
            )

    print("[INFO] Filtering log phase (readiness completeness/count checks):")
    if invalid_rows:
        invalid_df = pd.DataFrame(invalid_rows)
        print(invalid_df.to_string(index=False))
        print(f"[INFO] Excluded {len(invalid_rows)} runs before pod/qps and NUMBER_OF_RUNS filtering.")
    else:
        print("[INFO] No runs excluded by readiness filtering.")

    return raw_df.loc[keep_mask].reset_index(drop=True)


def limit_runs_per_case(raw_df: pd.DataFrame, number_of_runs: int) -> pd.DataFrame:
    if number_of_runs <= 0 or raw_df.empty:
        return raw_df

    tmp = raw_df.copy()
    if "scheduler" in tmp.columns and tmp["scheduler"].notna().any():
        tmp["_case_scheduler"] = tmp["scheduler"].map(_norm_sched_name)
    else:
        tmp["_case_scheduler"] = tmp["system"].map(_norm_sched_name)

    limited = (
        tmp.sort_values(["_case_scheduler", "qps", "pod_count", "input_log", "run_id", "run_seq"], kind="stable")
        .groupby(["_case_scheduler", "qps", "pod_count"], dropna=False, group_keys=False)
        .head(number_of_runs)
        .drop(columns=["_case_scheduler"])
        .reset_index(drop=True)
    )
    dropped = len(raw_df) - len(limited)
    if dropped:
        print(f"[INFO] Applied NUMBER_OF_RUNS={number_of_runs}: dropped {dropped} extra runs.")
    else:
        print(f"[INFO] Applied NUMBER_OF_RUNS={number_of_runs}: no extra runs found.")
    return limited


def add_resource_derived_metrics(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    rx_col = "ResourceUsage_net_rx_bytes_total"
    tx_col = "ResourceUsage_net_tx_bytes_total"
    if rx_col in out.columns and tx_col in out.columns:
        rx = pd.to_numeric(out[rx_col], errors="coerce")
        tx = pd.to_numeric(out[tx_col], errors="coerce")
        out["ResourceUsage_net_total_bytes"] = rx + tx
    return out


def build_excel_with_summary_and_charts(raw_df: pd.DataFrame, out_xlsx: Path) -> None:
    raw_df = add_resource_derived_metrics(raw_df)
    raw_x = to_nullable_ints(raw_df)
    summary = add_resource_derived_metrics(summarize_means(raw_df))
    raw_x = raw_x.astype(object).where(pd.notna(raw_x), None)
    summary = summary.astype(object).where(pd.notna(summary), None)

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw"
    ws_sum = wb.create_sheet("Summary_means")
    ws_ch = wb.create_sheet("Charts")

    for r in dataframe_to_rows(raw_x, index=False, header=True):
        ws_raw.append(r)
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws_sum.append(r)

    pod_counts = sorted(int(x) for x in summary["pod_count"].dropna().unique().tolist())
    qps_values = sorted(int(x) for x in summary["qps"].dropna().unique().tolist())
    systems = _ordered_systems(summary)

    metrics = {
        "PodScheduled": (
            "PodScheduled_min",
            "PodScheduled_q1",
            "PodScheduled_p50",
            "PodScheduled_q3",
            "PodScheduled_max",
            "PodScheduled_avg",
        ),
        "ContainerReady": (
            "ContainerReady_min",
            "ContainerReady_q1",
            "ContainerReady_p50",
            "ContainerReady_q3",
            "ContainerReady_max",
            "ContainerReady_avg",
        ),
        "ServiceLatency": ("ServiceLatency_avg", "ServiceLatency_p50", "ServiceLatency_99th", "ServiceLatency_max"),
    }
    resource_metrics = {
        "ResourceUsage_runningPods": "ResourceUsage_runningPods",
        "ResourceUsage_cpu_total_m": "ResourceUsage_cpu_total_m",
        "ResourceUsage_mem_total_mi": "ResourceUsage_mem_total_mi",
        "ResourceUsage_net_rx_bytes_total": "ResourceUsage_net_rx_bytes_total",
        "ResourceUsage_net_tx_bytes_total": "ResourceUsage_net_tx_bytes_total",
        "ResourceUsage_net_total_bytes": "ResourceUsage_net_total_bytes",
        "ContainerReady_expectedContainers": "ContainerReady_expectedContainers",
        "ContainerReady_readyContainers": "ContainerReady_readyContainers",
    }
    metric_stats = {
        "PodScheduled": [("Min", 0), ("Q1", 1), ("Median", 2), ("Q3", 3), ("Max", 4), ("Avg", 5)],
        "ContainerReady": [("Min", 0), ("Q1", 1), ("Median", 2), ("Q3", 3), ("Max", 4), ("Avg", 5)],
        "ServiceLatency": [("Avg", 0), ("P50", 1), ("P99", 2), ("Max", 3)],
    }

    def write_table(start_row: int, start_col: int, qps: int, metric_name: str, stat_label: str, stat_col: str):
        ws_ch.cell(row=start_row, column=start_col, value=f"QPS {qps} - {metric_name} ({stat_label})")
        ws_ch.cell(row=start_row + 1, column=start_col, value="pod_count")
        for j, sys in enumerate(systems):
            ws_ch.cell(row=start_row + 1, column=start_col + 1 + j, value=sys)

        for i, pods in enumerate(pod_counts):
            r = start_row + 2 + i
            ws_ch.cell(row=r, column=start_col, value=pods)

            for j, sys in enumerate(systems):
                v_s = summary.loc[
                    (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), stat_col
                ]
                s_val = v_s.iloc[0] if len(v_s) else None
                ws_ch.cell(row=r, column=start_col + 1 + j, value=float(s_val) if pd.notna(s_val) else None)

        header_row = start_row + 1
        last_row = start_row + 1 + len(pod_counts)
        return header_row, last_row

    def add_chart(
        table_header_row: int,
        table_left_col: int,
        table_last_row: int,
        title: str,
        anchor_cell: str,
        chart_kind: str,
    ):
        chart = BarChart() if chart_kind == "bar" else LineChart()
        chart.title = title
        chart.x_axis.title = "number of pods"
        chart.y_axis.title = "value (ms)"
        chart.legend.position = "r"
        if chart_kind == "bar":
            chart.type = "col"
            chart.style = 10

        cats = Reference(ws_ch, min_col=table_left_col, min_row=table_header_row + 1, max_row=table_last_row)
        data = Reference(
            ws_ch,
            min_col=table_left_col + 1,
            min_row=table_header_row,
            max_col=table_left_col + len(systems),
            max_row=table_last_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = False
        ws_ch.add_chart(chart, anchor_cell)

    row_cursor = 1
    left_col = 1
    for qps in qps_values:
        ws_ch.cell(row=row_cursor, column=left_col, value=f"QPS {qps} (QoS vs Default across number of pods)")
        row_cursor += 1
        for metric_name, cols in metrics.items():
            for stat_label, idx in metric_stats[metric_name]:
                stat_col = cols[idx]
                top = row_cursor
                header_row, last_row = write_table(top, left_col, qps, metric_name, stat_label, stat_col)
                anchor = ws_ch.cell(row=top, column=left_col + 4).coordinate
                chart_kind = "bar" if stat_label in {"Median", "P50", "Avg"} else "line"
                add_chart(header_row, left_col, last_row, f"QPS {qps} - {metric_name} ({stat_label})", anchor, chart_kind)
                row_cursor = last_row + 4
        for metric_name, stat_col in resource_metrics.items():
            top = row_cursor
            header_row, last_row = write_table(top, left_col, qps, metric_name, "Total", stat_col)
            anchor = ws_ch.cell(row=top, column=left_col + 4).coordinate
            add_chart(header_row, left_col, last_row, f"QPS {qps} - {metric_name} (Total)", anchor, "bar")
            row_cursor = last_row + 4
        row_cursor += 2

    for ws in [ws_raw, ws_sum, ws_ch]:
        max_col = min(ws.max_column, 22)
        for col_idx in range(1, max_col + 1):
            col_letter = chr(64 + col_idx)
            max_len = 0
            for row_idx in range(1, min(ws.max_row, 200) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(out_xlsx)


def _auto_log_scale(values: np.ndarray) -> bool:
    vals = values[np.isfinite(values) & (values > 0)]
    if len(vals) < 2:
        return False
    return (np.nanmax(vals) / np.nanmin(vals)) >= 50


def _place_legend_outside(ax) -> bool:
    handles, labels = ax.get_legend_handles_labels()
    if not handles:
        return False
    ax.legend(loc="upper left", bbox_to_anchor=(1.02, 1.0), borderaxespad=0.0)
    return True


def _save_chart_with_outside_legend(fig, ax, out_path: Path, dpi: int) -> None:
    has_legend = _place_legend_outside(ax)
    fig.tight_layout(rect=(0, 0, 0.82, 1) if has_legend else None)
    fig.savefig(out_path, dpi=dpi)


def plot_boxrange_xpods2(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)
    systems = _ordered_systems(summary)

    metric_map = {
        "PodScheduled": (
            "PodScheduled_min",
            "PodScheduled_q1",
            "PodScheduled_p50",
            "PodScheduled_q3",
            "PodScheduled_avg",
            "PodScheduled_max",
        ),
        "ContainerReady": (
            "ContainerReady_min",
            "ContainerReady_q1",
            "ContainerReady_p50",
            "ContainerReady_q3",
            "ContainerReady_avg",
            "ContainerReady_max",
        ),
        "ServiceLatency": (
            "ServiceLatency_p50",
            "ServiceLatency_p50",
            "ServiceLatency_p50",
            "ServiceLatency_99th",
            "ServiceLatency_avg",
            "ServiceLatency_max",
        ),
    }

    created: List[Path] = []
    box_width = 0.22

    for qps in qps_values:
        for metric_name, (min_col, q1_col, p50_col, q3_col, avg_col, max_col) in metric_map.items():
            vals = []
            for sys in systems:
                sub = summary[
                    (summary["system"] == sys)
                    & (summary["qps"] == qps)
                    & (summary["pod_count"].isin(pod_counts))
                ]
                if not sub.empty:
                    vals.extend(sub[[min_col, q1_col, p50_col, q3_col, avg_col, max_col]].to_numpy().ravel().tolist())

            all_vals = np.array([v for v in vals if v is not None and pd.notna(v)], dtype=float)
            use_log = _auto_log_scale(all_vals) if len(all_vals) else False
            if use_log and not np.all(all_vals[np.isfinite(all_vals)] > 0):
                use_log = False

            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("Time (ms)")
            ax.set_title(f"QPS {qps} - {metric_name}")

            nsys = max(len(systems), 1)
            if nsys == 1:
                offsets = [0.0]
            else:
                span = 0.36
                step = span / (nsys - 1)
                offsets = [(-span / 2) + i * step for i in range(nsys)]

            legend_done = set()

            for idx, sys in enumerate(systems):
                dx = offsets[idx]
                color = _system_color(sys)
                label = _display_system(sys)

                sub = summary[
                    (summary["system"] == sys)
                    & (summary["qps"] == qps)
                    & (summary["pod_count"].isin(pod_counts))
                ].copy()

                sub["pod_count"] = pd.Categorical(sub["pod_count"], categories=list(pod_counts), ordered=True)
                sub = sub.sort_values("pod_count")

                for i, pods in enumerate(pod_counts):
                    row = sub[sub["pod_count"] == pods]
                    if row.empty:
                        continue

                    mn = row.iloc[0][min_col]
                    q1 = row.iloc[0][q1_col]
                    p50 = row.iloc[0][p50_col]
                    q3 = row.iloc[0][q3_col]
                    avg = row.iloc[0][avg_col]
                    mx = row.iloc[0][max_col]

                    if pd.isna(mn) and pd.isna(q1) and pd.isna(p50) and pd.isna(q3) and pd.isna(avg) and pd.isna(mx):
                        continue

                    xi = x[i] + dx

                    mnf = float(mn) if pd.notna(mn) else np.nan
                    q1f = float(q1) if pd.notna(q1) else np.nan
                    p50f = float(p50) if pd.notna(p50) else np.nan
                    q3f = float(q3) if pd.notna(q3) else np.nan
                    avgf = float(avg) if pd.notna(avg) else np.nan
                    mxf = float(mx) if pd.notna(mx) else np.nan

                    if np.isfinite(q1f) and np.isfinite(q3f) and q3f >= q1f:
                        rect = Rectangle(
                            (xi - box_width / 2, q1f),
                            box_width,
                            q3f - q1f,
                            facecolor="none",
                            edgecolor=color,
                            linewidth=1.8,
                            label=label if label not in legend_done else None,
                        )
                        ax.add_patch(rect)
                        legend_done.add(label)

                    if np.isfinite(p50f):
                        ax.hlines(
                            p50f,
                            xi - box_width / 2,
                            xi + box_width / 2,
                            colors=color,
                            linewidth=2.0,
                        )

                    if np.isfinite(avgf):
                        ax.plot([xi], [avgf], marker="o", linestyle="None", color=color, markersize=4)

                    if np.isfinite(mnf) and np.isfinite(q1f) and q1f >= mnf:
                        ax.vlines(xi, mnf, q1f, colors=color, linewidth=1.5)
                        ax.hlines(
                            mnf,
                            xi - box_width * 0.22,
                            xi + box_width * 0.22,
                            colors=color,
                            linewidth=1.5,
                        )

                    if np.isfinite(q3f) and np.isfinite(mxf) and mxf >= q3f:
                        ax.vlines(xi, q3f, mxf, colors=color, linewidth=1.5)
                        ax.hlines(
                            mxf,
                            xi - box_width * 0.22,
                            xi + box_width * 0.22,
                            colors=color,
                            linewidth=1.5,
                        )

            ax.grid(True, which="both", axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            if use_log:
                ax.set_yscale("log")

            out_path = out_dir / f"boxrange2_{metric_name}_qps{qps}.png"
            _save_chart_with_outside_legend(fig, ax, out_path, dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_boxrange_xpods(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
    label_avg: bool = False,
    avg_label_decimals: int = 0,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)
    systems = _ordered_systems(summary)

    metric_map = {
        "PodScheduled": ("PodScheduled_avg", "PodScheduled_p50", "PodScheduled_p99", "PodScheduled_max"),
        "ContainerReady": ("ContainerReady_avg", "ContainerReady_p50", "ContainerReady_p99", "ContainerReady_max"),
        "ServiceLatency": ("ServiceLatency_avg", "ServiceLatency_p50", "ServiceLatency_99th", "ServiceLatency_max"),
    }
    box_width = 0.30
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, (avg_col, p50_col, p99_col, max_col) in metric_map.items():
            vals = []
            for sys in systems:
                sub = summary[
                    (summary["system"] == sys)
                    & (summary["qps"] == qps)
                    & (summary["pod_count"].isin(pod_counts))
                ]
                if not sub.empty:
                    vals.extend(sub[[avg_col, p50_col, p99_col, max_col]].to_numpy().ravel().tolist())
            all_vals = np.array([v for v in vals if v is not None], dtype=float)
            use_log = _auto_log_scale(all_vals)
            if use_log and not np.all(all_vals[np.isfinite(all_vals)] > 0):
                use_log = False

            fig, ax = plt.subplots(figsize=(7.8, 4.6))
            x = np.arange(len(pod_counts), dtype=float)
            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("Time (ms)")
            ax.set_title(f"QPS {qps} - {metric_name} (box=p99..max, dot=avg)")

            nsys = max(len(systems), 1)
            if nsys == 1:
                offsets = [0.0]
            else:
                span = 0.54
                step = span / (nsys - 1)
                offsets = [(-span / 2) + i * step for i in range(nsys)]

            for idx, sys in enumerate(systems):
                dx = offsets[idx]
                label = _display_system(sys)
                color = _system_color(sys)
                sub = summary[
                    (summary["system"] == sys)
                    & (summary["qps"] == qps)
                    & (summary["pod_count"].isin(pod_counts))
                ].copy()
                sub["pod_count"] = pd.Categorical(sub["pod_count"], categories=list(pod_counts), ordered=True)
                sub = sub.sort_values("pod_count")

                proxy_done = False
                for i, pods in enumerate(pod_counts):
                    row = sub[sub["pod_count"] == pods]
                    if row.empty:
                        continue
                    avg = row.iloc[0][avg_col]
                    p99 = row.iloc[0][p99_col]
                    mx = row.iloc[0][max_col]
                    if pd.isna(avg) and pd.isna(p99) and pd.isna(mx):
                        continue

                    xi = x[i] + dx
                    if pd.notna(p99) and pd.notna(mx) and float(mx) >= float(p99):
                        rect = Rectangle(
                            (xi - box_width / 2, float(p99)),
                            box_width,
                            float(mx) - float(p99),
                            fill=False,
                            linewidth=1.8,
                            edgecolor=color,
                            label=label if not proxy_done else None,
                        )
                        ax.add_patch(rect)
                        proxy_done = True

                    if pd.notna(avg):
                        y = float(avg)
                        ax.plot([xi], [y], marker="o", linestyle="None", color=color)
                        if label_avg:
                            fmt = f"{{:.{avg_label_decimals}f}}"
                            voff = 6 if idx % 2 == 0 else -10
                            ax.annotate(
                                fmt.format(y),
                                (xi, y),
                                textcoords="offset points",
                                xytext=(0, voff),
                                ha="center",
                                va="bottom" if voff > 0 else "top",
                                color=color,
                                fontsize=8,
                            )

            ax.grid(True, which="both", axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            if use_log:
                ax.set_yscale("log")

            out_path = out_dir / f"boxrange_{metric_name}_qps{qps}.png"
            _save_chart_with_outside_legend(fig, ax, out_path, dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_p50_bar_xpods(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)
    systems = _ordered_systems(summary)

    metric_map = {
        "PodScheduled": "PodScheduled_p50",
        "ContainerReady": "ContainerReady_p50",
        "ServiceLatency": "ServiceLatency_p50",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, p50_col in metric_map.items():
            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), p50_col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("P50 (ms)")
            ax.set_title(f"QPS {qps} - {metric_name} (P50 bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)

            out_path = out_dir / f"bar_p50_{metric_name}_qps{qps}.png"
            _save_chart_with_outside_legend(fig, ax, out_path, dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_avg_bar_xpods(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)
    systems = _ordered_systems(summary)

    metric_map = {
        "PodScheduled": "PodScheduled_avg",
        "ContainerReady": "ContainerReady_avg",
        "ServiceLatency": "ServiceLatency_avg",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, avg_col in metric_map.items():
            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), avg_col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("AVG (ms)")
            ax.set_title(f"QPS {qps} - {metric_name} (AVG bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)

            out_path = out_dir / f"bar_avg_{metric_name}_qps{qps}.png"
            _save_chart_with_outside_legend(fig, ax, out_path, dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_resource_bar_xpods(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = add_resource_derived_metrics(summarize_means(add_resource_derived_metrics(raw_df)))
    systems = _ordered_systems(summary)

    metric_map = {
        "ResourceUsage_runningPods": "ResourceUsage_runningPods",
        "ResourceUsage_cpu_total_m": "ResourceUsage_cpu_total_m",
        "ResourceUsage_mem_total_mi": "ResourceUsage_mem_total_mi",
        "ResourceUsage_net_rx_bytes_total": "ResourceUsage_net_rx_bytes_total",
        "ResourceUsage_net_tx_bytes_total": "ResourceUsage_net_tx_bytes_total",
        "ResourceUsage_net_total_bytes": "ResourceUsage_net_total_bytes",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, col in metric_map.items():
            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel(metric_name)
            ax.set_title(f"QPS {qps} - {metric_name} (bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)

            out_path = out_dir / f"bar_resource_{metric_name}_qps{qps}.png"
            _save_chart_with_outside_legend(fig, ax, out_path, dpi)
            plt.close(fig)
            created.append(out_path)
    return created


def plot_resource_bar_xpods2(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = add_resource_derived_metrics(summarize_means(add_resource_derived_metrics(raw_df)))
    systems = _ordered_systems(summary)

    metric_map = {
        "ResourceUsage_runningPods": ("running pods", "count"),
        "ResourceUsage_cpu_total_m": ("total CPU", "millicores"),
        "ResourceUsage_mem_total_mi": ("total memory", "MiB"),
        "ResourceUsage_net_rx_bytes_total": ("total RX bytes", "bytes"),
        "ResourceUsage_net_tx_bytes_total": ("total TX bytes", "bytes"),
        "ResourceUsage_net_total_bytes": ("total network bytes", "bytes"),
    }

    created: List[Path] = []

    for qps in qps_values:
        for metric_name, (display_name, ylab) in metric_map.items():
            vals_all = []
            for sys in systems:
                sub = summary[
                    (summary["system"] == sys)
                    & (summary["qps"] == qps)
                    & (summary["pod_count"].isin(pod_counts))
                ]
                if not sub.empty and metric_name in sub.columns:
                    vals_all.extend(sub[metric_name].dropna().tolist())

            all_vals = np.array([v for v in vals_all if v is not None and pd.notna(v)], dtype=float)
            use_log = _auto_log_scale(all_vals) if len(all_vals) else False
            if use_log and not np.all(all_vals[np.isfinite(all_vals)] > 0):
                use_log = False

            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)

            nsys = max(len(systems), 1)
            group_w = 0.78
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys)
                        & (summary["qps"] == qps)
                        & (summary["pod_count"] == pods),
                        metric_name,
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)

                xpos = x + start + idx * width
                ax.bar(
                    xpos,
                    vals,
                    width=width * 0.92,
                    color=_system_color(sys),
                    label=_display_system(sys),
                    alpha=0.9,
                )

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel(ylab)
            ax.set_title(f"QPS {qps} - {display_name}")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)

            if use_log:
                ax.set_yscale("log")

            out_path = out_dir / f"bar_resource2_{metric_name}_qps{qps}.png"
            _save_chart_with_outside_legend(fig, ax, out_path, dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_containerready_counts_bar_xpods(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)
    systems = _ordered_systems(summary)

    metric_map = {
        "readyContainers": "ContainerReady_readyContainers",
        "expectedContainers": "ContainerReady_expectedContainers",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, col in metric_map.items():
            if col not in summary.columns:
                continue

            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("containers")
            ax.set_title(f"QPS {qps} - ContainerReady {metric_name} (bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)

            out_path = out_dir / f"bar_containerready_{metric_name}_qps{qps}.png"
            _save_chart_with_outside_legend(fig, ax, out_path, dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def merge_metric_qps_plots(
    created_paths: List[Path],
    out_dir: Path,
    parse_regex: str,
    out_name_tpl: str,
    title_tpl: str,
    ncols: int = 2,
    dpi: int = 200,
    group_by: str = "metric",
    max_panels_per_figure: int = 4,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    rx = re.compile(parse_regex)
    grouped: dict[str, list[tuple[str, Path]]] = {}

    for p in created_paths:
        m = rx.match(p.name)
        if not m:
            continue
        metric = m.group("metric")
        qps = int(m.group("qps"))
        if group_by == "metric":
            grouped.setdefault(metric, []).append((qps, p))
        elif group_by == "qps":
            grouped.setdefault(str(qps), []).append((metric, p))

    merged: List[Path] = []
    for key, items in grouped.items():
        items = sorted(items, key=lambda t: t[0])
        n = len(items)
        if n == 0:
            continue

        chunk_size = max(1, max_panels_per_figure)
        pages = math.ceil(n / chunk_size)
        for page in range(pages):
            part = items[page * chunk_size : (page + 1) * chunk_size]
            pn = len(part)
            nrows = math.ceil(pn / ncols)
            fig, axes = plt.subplots(nrows, ncols, figsize=(6.4 * ncols, 4.6 * nrows))
            axes_arr = np.array(axes).reshape(-1)

            for idx, (subkey, pth) in enumerate(part):
                ax = axes_arr[idx]
                ax.imshow(plt.imread(pth))
                ax.axis("off")
                panel = chr(ord("a") + idx) if idx < 26 else str(idx + 1)
                label_text = f"({panel}) QPS {subkey}" if group_by == "metric" else f"({panel}) {subkey}"
                ax.text(
                    0.5,
                    -0.12,
                    label_text,
                    transform=ax.transAxes,
                    ha="center",
                    va="top",
                    fontsize=11,
                    clip_on=False,
                )

            for idx in range(pn, len(axes_arr)):
                axes_arr[idx].axis("off")

            title = title_tpl.format(**{group_by: key})
            if pages > 1:
                title = f"{title} (part {page + 1}/{pages})"
            fig.suptitle(title, fontsize=14, y=0.98)
            bottom = 0.15 if nrows == 1 else 0.06
            fig.subplots_adjust(left=0.04, right=0.98, bottom=bottom, top=0.90, wspace=0.04, hspace=0.35)

            base_name = out_name_tpl.format(**{group_by: key})
            if pages > 1:
                stem = Path(base_name).stem
                suffix = Path(base_name).suffix
                out_name = f"{stem}_part{page + 1}{suffix}"
            else:
                out_name = base_name
            out_path = out_dir / out_name
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            gc.collect()
            merged.append(out_path)

    return merged


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parse bookinfo summaries and plot with pod_count as X-axis (split by qps and metric)."
    )
    parser.add_argument("min_pod", nargs="?", type=int, help="Optional minimum pod count to include.")
    parser.add_argument("max_pod", nargs="?", type=int, help="Optional maximum pod count to include.")
    parser.add_argument("min_qps", nargs="?", type=int, help="Optional minimum qps to include.")
    parser.add_argument("max_qps", nargs="?", type=int, help="Optional maximum qps to include.")
    parser.add_argument(
        "--base-dir",
        default="/home/samizadeh/Downloads/swm-git/swm-benchmarking-paper-main/Experiment1/logs/",
        help="Directory containing log files and where outputs are written.",
    )
    parser.add_argument("--glob", default="*pod-summary_*.log", help="Glob pattern to select input logs.")
    parser.add_argument(
        "--out-excel",
        default="pod_summary_qos_vs_def_xpods.xlsx",
        help="Output Excel filename (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--plots-dir",
        default="plots_boxrange_xpods",
        help="Directory for PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--plots-dir2",
        default="plots_boxrange_xpods2",
        help="Directory for alternative box-style latency PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--p50-plots-dir",
        default="plots_p50_bar_xpods",
        help="Directory for additional P50 bar PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--avg-plots-dir",
        default="plots_avg_bar_xpods",
        help="Directory for additional AVG bar PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--resource-plots-dir",
        default="plots_resource_bar_xpods",
        help="Directory for ResourceUsage bar PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--resource-plots-dir2",
        default="plots_resource_bar_xpods2",
        help="Directory for alternative ResourceUsage bar PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--containerready-plots-dir",
        default="plots_containerready_bar_xpods",
        help="Directory for ContainerReady count bar PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument("--pods", default="", help="Optional comma-separated pod counts (e.g. '10,20,40').")
    parser.add_argument("--qps", default="", help="Optional comma-separated qps values (e.g. '10,100,500').")
    parser.add_argument("--dpi", type=int, default=120, help="Output DPI for generated plots (lower uses less memory).")
    parser.add_argument(
        "--merge-max-panels",
        type=int,
        default=4,
        help="Maximum number of single charts loaded into one merged panel image.",
    )
    parser.add_argument("--label-avg", action="store_true", help="Write avg value label near each avg dot.")
    parser.add_argument("--avg-label-decimals", type=int, default=0, help="Decimals for avg labels.")
    single_group = parser.add_mutually_exclusive_group()
    single_group.add_argument(
        "--single-charts",
        dest="single_charts",
        action="store_true",
        help="Keep single PNG charts in addition to merged panel charts.",
    )
    single_group.add_argument(
        "--no-single-charts",
        dest="single_charts",
        action="store_false",
        help="Do not keep single PNG charts; only keep merged panel charts.",
    )
    parser.set_defaults(single_charts=False)
    args = parser.parse_args()
    if (args.min_pod is None) != (args.max_pod is None):
        raise SystemExit("ERROR: provide both min_pod and max_pod, e.g. parse_and_plot_pod_summary_xpods.py 5 15")
    if args.min_pod is not None and args.max_pod < args.min_pod:
        raise SystemExit("ERROR: max_pod must be greater than or equal to min_pod.")
    if args.min_qps is not None and args.min_pod is None:
        raise SystemExit("ERROR: provide pod range before qps range, e.g. parse_and_plot_pod_summary_xpods.py 5 15 5 20")
    if (args.min_qps is None) != (args.max_qps is None):
        raise SystemExit("ERROR: provide both min_qps and max_qps, e.g. parse_and_plot_pod_summary_xpods.py 5 15 5 20")
    if args.min_qps is not None and args.max_qps < args.min_qps:
        raise SystemExit("ERROR: max_qps must be greater than or equal to min_qps.")

    base_dir = Path(args.base_dir).expanduser()
    if not base_dir.exists():
        raise SystemExit(f"ERROR: base dir does not exist: {base_dir}")

    def resolve(p: str) -> Path:
        pp = Path(p).expanduser()
        return pp if pp.is_absolute() else base_dir / pp

    log_paths = sorted(base_dir.glob(args.glob))
    if not log_paths:
        raise SystemExit(f"ERROR: no logs matched '{args.glob}' under {base_dir}")
    log_paths = filter_log_paths_by_readiness(log_paths)
    if not log_paths:
        raise SystemExit("ERROR: no log files left after file readiness filtering.")

    raw_parts = [parse_file(p) for p in log_paths]
    raw_df = pd.concat(raw_parts, ignore_index=True)
    if raw_df.empty:
        raise SystemExit("ERROR: no parseable metric blocks found in input logs.")
    raw_df = add_pod_scheduled_counts(raw_df, log_paths)
    raw_df = filter_readiness_runs(raw_df)
    if raw_df.empty:
        raise SystemExit("ERROR: no complete runs left after readiness filtering.")
    if args.min_pod is not None:
        raw_df = raw_df[
            pd.to_numeric(raw_df["pod_count"], errors="coerce").between(args.min_pod, args.max_pod, inclusive="both")
        ].reset_index(drop=True)
        if raw_df.empty:
            raise SystemExit(f"ERROR: no rows left after pod range filtering: {args.min_pod}..{args.max_pod}")
        print(f"[INFO] Applied pod range filter: {args.min_pod}..{args.max_pod}")
    if args.min_qps is not None:
        raw_df = raw_df[
            pd.to_numeric(raw_df["qps"], errors="coerce").between(args.min_qps, args.max_qps, inclusive="both")
        ].reset_index(drop=True)
        if raw_df.empty:
            raise SystemExit(f"ERROR: no rows left after qps range filtering: {args.min_qps}..{args.max_qps}")
        print(f"[INFO] Applied qps range filter: {args.min_qps}..{args.max_qps}")
    print_run_count_table(raw_df, "actual after filters, before NUMBER_OF_RUNS")
    print("[INFO] ------------------------------------------------------------")
    raw_df = limit_runs_per_case(raw_df, NUMBER_OF_RUNS)

    print_run_count_table(raw_df, "selected for charts after NUMBER_OF_RUNS")

    out_excel = resolve(args.out_excel)
    build_excel_with_summary_and_charts(raw_df, out_excel)
    print(f"[OK] Excel written: {out_excel}")

    if args.pods.strip():
        pod_counts = tuple(int(x.strip()) for x in args.pods.split(",") if x.strip())
    else:
        pod_counts = tuple(sorted(int(x) for x in raw_df["pod_count"].dropna().unique()))

    if args.qps.strip():
        qps_values = tuple(int(x.strip()) for x in args.qps.split(",") if x.strip())
    else:
        qps_values = tuple(sorted(int(x) for x in raw_df["qps"].dropna().unique()))

    plots_dir2 = resolve(args.plots_dir2)
    created2 = plot_boxrange_xpods2(
        raw_df=raw_df,
        out_dir=plots_dir2,
        pod_counts=pod_counts,
        qps_values=qps_values,
        dpi=args.dpi,
    )
    print(f"[OK] {len(created2)} alternative boxrange plots written to: {plots_dir2}")
    boxrange2_panel_inputs = [p for p in created2 if not p.name.startswith("boxrange2_ServiceLatency_")]
    merged_box2 = merge_metric_qps_plots(
        created_paths=boxrange2_panel_inputs,
        out_dir=plots_dir2,
        parse_regex=r"^boxrange2_(?P<metric>.+)_qps(?P<qps>\d+)\.png$",
        out_name_tpl="panel_2x2_boxrange2_{metric}.png",
        title_tpl="{metric} - Summary box plots across QPS",
        dpi=args.dpi,
        max_panels_per_figure=args.merge_max_panels,
    )
    print(f"[OK] {len(merged_box2)} merged alternative boxrange panels written to: {plots_dir2}")
    if not args.single_charts:
        removed = 0
        for p in created2:
            if p.exists():
                p.unlink()
                removed += 1
        print(f"[OK] Removed {removed} single alternative boxrange plots (kept merged panels only).")
    gc.collect()

    resource_plots_dir = resolve(args.resource_plots_dir)
    created_resource = plot_resource_bar_xpods(
        raw_df=raw_df,
        out_dir=resource_plots_dir,
        pod_counts=pod_counts,
        qps_values=qps_values,
        dpi=args.dpi,
    )
    print(f"[OK] {len(created_resource)} ResourceUsage bar plots written to: {resource_plots_dir}")
    merged_resource = merge_metric_qps_plots(
        created_paths=created_resource,
        out_dir=resource_plots_dir,
        parse_regex=r"^bar_resource_(?P<metric>.*?)_qps(?P<qps>\d+)\.png$",
        out_name_tpl="panel_2x2_bar_resource_{metric}.png",
        title_tpl="Resource Usage: {metric}",
        ncols=2,
        dpi=args.dpi,
        max_panels_per_figure=args.merge_max_panels,
    )
    print(f"[OK] {len(merged_resource)} merged resource panels written to: {resource_plots_dir}")
    if not args.single_charts:
        removed = 0
        for p in created_resource:
            if p.exists():
                p.unlink()
                removed += 1
        print(f"[OK] Removed {removed} single resource bar plots (kept merged panels only).")
    gc.collect()


if __name__ == "__main__":
    main()
