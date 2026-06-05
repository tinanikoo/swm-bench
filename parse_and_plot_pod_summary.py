#!/usr/bin/env python3
"""
Parse bookinfo pod summary logs, build Excel report, and create PNG plots.

This parser is designed for logs like:
  pod-summary_*_qos-scheduler_20.log
  pod-summary_*_default-scheduler_20.log

Key grouping dimensions:
  - pod_count (extracted from filename suffix, e.g. _20.log)
  - qps
  - scheduler/system (qos vs def)

Output:
  - Excel workbook with raw rows, summary means, and line charts
  - Python PNG box-range charts (box: p99..max, dot: avg)
"""

from __future__ import annotations

import argparse
import gc
import math
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.patches import Rectangle
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows


FILE_POD_RE = re.compile(r"_(?P<pod_count>\d+)\.log$")
SCHEDULER_SUFFIX_RE = re.compile(r"_(?:qos(?:-assured)?|default|def)-scheduler_(?P<tail>.+)\.log$", flags=re.IGNORECASE)
SCHEDULER_HEADER_RE = re.compile(r"^#\s*scheduler:\s*(?P<scheduler>[^\n]+)\s*$", flags=re.MULTILINE)
METRIC_LINE_RE = re.compile(
    r"(?P<metric>PodScheduledTime|ContainerReadyTime|ServiceLatency)\s+"
    r"run=(?P<run_id>\d+)\s+"
    r"(?:case_no=(?P<case_no>\d+)\s+)?"
    r"jobIterations=(?P<job_iterations>\d+)\s+"
    r"qps=(?P<qps>\d+)\s+"
    r"burst=(?P<burst>\d+)\s+"
    r"bookinfo_replicas=(?P<bookinfo_replicas>\d+)\s+"
    r"observed=(?P<observed>\S+)\s+"
    r"(?P<stats>[^\n]+)"
)
RESOURCE_LINE_RE = re.compile(
    r"ResourceUsage\s+run=(?P<run_id>\d+)\s+"
    r"jobIterations=(?P<job_iterations>\d+)\s+"
    r"qps=(?P<qps>\d+)\s+"
    r"burst=(?P<burst>\d+)\s+"
    r"bookinfo_replicas=(?P<bookinfo_replicas>\d+)\s+"
    r"(?:deployment_duration_s=(?P<deployment_duration_s>\d+)\s+)?"
    r"runningPods=(?P<running_pods>\d+)\s+"
    r"cpu_total_m=(?P<cpu_total_m>\d+|na)\s+"
    r"mem_total_mi=(?P<mem_total_mi>\d+|na)\s+"
    r"net_rx_bytes_total=(?P<net_rx_bytes_total>\d+|na)\s+"
    r"net_tx_bytes_total=(?P<net_tx_bytes_total>\d+|na)"
    r"(?:\s+net_status=(?P<net_status>\S+))?"
    r"(?:\s+net_pods_sampled=(?P<net_pods_sampled>\d+))?"
)
CONTAINER_READY_COUNTS_RE = re.compile(
    r"ContainerReadyTime[^\n]*\bexpectedContainers=(?P<expected_containers>\d+)\b[^\n]*\breadyContainers=(?P<ready_containers>\d+)\b"
)

RAW_COLS = [
    "input_log",
    "scheduler",
    "system",
    "pod_count",
    "run_id",
    "run_seq",
    "qps",
    "ResourceUsage_runningPods",
    "ResourceUsage_cpu_total_m",
    "ResourceUsage_mem_total_mi",
    "ResourceUsage_net_rx_bytes_total",
    "ResourceUsage_net_tx_bytes_total",
    "PodScheduled_min",
    "PodScheduled_q1",
    "PodScheduled_p50",
    "PodScheduled_q3",
    "PodScheduled_max",
    "PodScheduled_avg",
    "ContainerReady_min",
    "ContainerReady_q1",
    "ContainerReady_p50",
    "ContainerReady_q3",
    "ContainerReady_max",
    "ContainerReady_avg",
    "ContainerReady_expectedContainers",
    "ContainerReady_readyContainers",
    "ServiceLatency_99th",
    "ServiceLatency_p50",
    "ServiceLatency_max",
    "ServiceLatency_avg",
    # Backward-compatible aliases
    "PodScheduled_p99",
    "ContainerReady_p99",
]

METRIC_PREFIX = {
    "PodScheduledTime": "PodScheduled",
    "ContainerReadyTime": "ContainerReady",
    "ServiceLatency": "ServiceLatency",
}


def split_blocks(text: str) -> List[str]:
    return re.split(r"=+\n", text)


def scheduler_to_system(scheduler: str, filename: str) -> str:
    s = scheduler.lower()
    f = filename.lower()
    if "assured" in s or "assured" in f:
        return "qos_assured"
    if "qos" in s or "qos" in f:
        return "qos"
    if "default" in s or "def" in f:
        return "def"
    return scheduler.strip().lower().replace(" ", "_")


def _pod_candidates_from_filename(filename: str) -> List[int]:
    """
    Return numeric candidates for pod_count from the filename suffix.

    Supports both:
      - ..._default-scheduler_20.log
      - ..._default-scheduler_20_50.log  (pods + qps)
    """
    m = SCHEDULER_SUFFIX_RE.search(filename)
    if m:
        nums = [int(x) for x in re.findall(r"\d+", m.group("tail"))]
        if nums:
            return nums

    # Backward-compatible fallback: old single-number suffix.
    pod_match = FILE_POD_RE.search(filename)
    if pod_match:
        return [int(pod_match.group("pod_count"))]
    return []


def _pick_pod_count(candidates: List[int], qps: int) -> Optional[int]:
    if not candidates:
        return None
    non_qps = [x for x in candidates if x != qps]
    return non_qps[0] if non_qps else candidates[0]


def parse_file(path: Path) -> pd.DataFrame:
    text = path.read_text(encoding="utf-8", errors="replace")
    sched_match = SCHEDULER_HEADER_RE.search(text)
    scheduler = sched_match.group("scheduler").strip() if sched_match else "unknown"
    system = scheduler_to_system(scheduler, path.name)

    pod_candidates = _pod_candidates_from_filename(path.name)

    rows: List[Dict[str, object]] = []
    run_seq = 0

    for block in split_blocks(text):
        matches = list(METRIC_LINE_RE.finditer(block))
        if not matches:
            continue

        run_seq += 1
        run_id = int(matches[0].group("run_id"))
        qps = int(matches[0].group("qps"))

        row: Dict[str, object] = {
            "input_log": path.name,
            "scheduler": scheduler,
            "system": system,
            "pod_count": _pick_pod_count(pod_candidates, qps),
            "run_id": run_id,
            "run_seq": run_seq,
            "qps": qps,
        }

        for m in matches:
            metric = m.group("metric")
            prefix = METRIC_PREFIX[metric]
            stats_text = m.group("stats")
            kv = dict(re.findall(r"([A-Za-z0-9_]+)=([^\s]+)", stats_text))

            def _to_ms_int(key: str) -> Optional[int]:
                raw = kv.get(key)
                if raw is None:
                    return None
                raw = raw.strip()
                if raw.endswith("ms"):
                    raw = raw[:-2]
                if raw.lower() in {"na", "nan", ""}:
                    return None
                try:
                    return int(raw)
                except ValueError:
                    return None

            if metric == "ServiceLatency":
                row["ServiceLatency_99th"] = _to_ms_int("99th")
                row["ServiceLatency_p50"] = _to_ms_int("p50")
                row["ServiceLatency_max"] = _to_ms_int("max")
                row["ServiceLatency_avg"] = _to_ms_int("avg")
            else:
                row[f"{prefix}_min"] = _to_ms_int("min")
                row[f"{prefix}_q1"] = _to_ms_int("q1")
                row[f"{prefix}_p50"] = _to_ms_int("median") or _to_ms_int("p50")
                row[f"{prefix}_q3"] = _to_ms_int("q3")
                row[f"{prefix}_max"] = _to_ms_int("max")
                row[f"{prefix}_avg"] = _to_ms_int("avg")
                # Keep compatibility with existing plotting code that expects p99.
                row[f"{prefix}_p99"] = _to_ms_int("99th") or _to_ms_int("q3")

        cr_counts = CONTAINER_READY_COUNTS_RE.search(block)
        if cr_counts:
            row["ContainerReady_expectedContainers"] = int(cr_counts.group("expected_containers"))
            row["ContainerReady_readyContainers"] = int(cr_counts.group("ready_containers"))

        rm = RESOURCE_LINE_RE.search(block)
        if rm:
            running_pods = int(rm.group("running_pods"))
            cpu_total_m = int(rm.group("cpu_total_m")) if rm.group("cpu_total_m") != "na" else None
            mem_total_mi = int(rm.group("mem_total_mi")) if rm.group("mem_total_mi") != "na" else None
            net_rx_total = int(rm.group("net_rx_bytes_total")) if rm.group("net_rx_bytes_total") != "na" else None
            net_tx_total = int(rm.group("net_tx_bytes_total")) if rm.group("net_tx_bytes_total") != "na" else None
            net_status = (rm.group("net_status") or "").strip().lower()
            net_pods_sampled = int(rm.group("net_pods_sampled")) if rm.group("net_pods_sampled") else 0

            # Ignore empty/unavailable ResourceUsage snapshots, e.g. runningPods=0 and all totals 0.
            is_empty_snapshot = (
                running_pods == 0
                and cpu_total_m in {None, 0}
                and mem_total_mi in {None, 0}
                and net_rx_total in {None, 0}
                and net_tx_total in {None, 0}
                and net_pods_sampled == 0
                and net_status in {"", "na", "unavailable"}
            )
            if not is_empty_snapshot:
                row["ResourceUsage_runningPods"] = running_pods
                row["ResourceUsage_cpu_total_m"] = cpu_total_m
                row["ResourceUsage_mem_total_mi"] = mem_total_mi
                row["ResourceUsage_net_rx_bytes_total"] = net_rx_total
                row["ResourceUsage_net_tx_bytes_total"] = net_tx_total

        rows.append(row)

    df = pd.DataFrame(rows)
    if df.empty:
        return pd.DataFrame(columns=RAW_COLS)

    df = df.reindex(columns=RAW_COLS)
    return df.sort_values(["pod_count", "run_id", "run_seq", "qps"], kind="stable").reset_index(drop=True)


def to_nullable_ints(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c in out.columns:
        if c in {"input_log", "scheduler", "system"}:
            continue
        out[c] = pd.to_numeric(out[c], errors="coerce").astype("Int64")
    return out


METRIC_COLS = [c for c in RAW_COLS if c not in {"input_log", "scheduler", "system", "pod_count", "run_id", "run_seq", "qps"}]


def summarize_means(raw_df: pd.DataFrame) -> pd.DataFrame:
    tmp = raw_df.copy()
    for c in METRIC_COLS:
        if c in tmp.columns:
            tmp[c] = pd.to_numeric(tmp[c], errors="coerce")
    grouped = tmp.groupby(["system", "pod_count", "qps"], as_index=False)[METRIC_COLS].mean(numeric_only=True)
    grouped = grouped.reindex(columns=["system", "pod_count", "qps", *METRIC_COLS])
    return grouped.round(2).sort_values(["pod_count", "qps", "system"], kind="stable").reset_index(drop=True)


def build_excel_with_summary_and_charts(raw_df: pd.DataFrame, out_xlsx: Path) -> None:
    raw_x = to_nullable_ints(raw_df)
    summary = summarize_means(raw_df)
    raw_x = raw_x.astype(object).where(pd.notna(raw_x), None)
    summary = summary.astype(object).where(pd.notna(summary), None)

    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw"
    ws_sum = wb.create_sheet("Summary_means")
    ws_ch = wb.create_sheet("Charts")

    for r in dataframe_to_rows(raw_x, index=False, header=True):
        ws_raw.append(r)
    for r in dataframe_to_rows(summary, index=False, header=True):
        ws_sum.append(r)

    pod_counts = sorted(int(x) for x in summary["pod_count"].dropna().unique().tolist())
    qps_values = sorted(int(x) for x in summary["qps"].dropna().unique().tolist())

    metrics = {
        "PodScheduled": ("PodScheduled_avg", "PodScheduled_p50", "PodScheduled_p99", "PodScheduled_max"),
        "ContainerReady": ("ContainerReady_avg", "ContainerReady_p50", "ContainerReady_p99", "ContainerReady_max"),
        "ServiceLatency": ("ServiceLatency_avg", "ServiceLatency_p50", "ServiceLatency_99th", "ServiceLatency_max"),
    }
    stats = [("Avg", 0), ("P50", 1), ("P99", 2), ("Max", 3)]

    def write_table(start_row: int, start_col: int, pod_count: int, metric_name: str, stat_label: str, stat_col: str):
        ws_ch.cell(row=start_row, column=start_col, value=f"Pods {pod_count} - {metric_name} ({stat_label})")
        ws_ch.cell(row=start_row + 1, column=start_col, value="qps")
        ws_ch.cell(row=start_row + 1, column=start_col + 1, value="qos")
        ws_ch.cell(row=start_row + 1, column=start_col + 2, value="def")

        for i, qps in enumerate(qps_values):
            r = start_row + 2 + i
            ws_ch.cell(row=r, column=start_col, value=qps)

            v_q = summary.loc[
                (summary["system"] == "qos") & (summary["pod_count"] == pod_count) & (summary["qps"] == qps), stat_col
            ]
            v_d = summary.loc[
                (summary["system"] == "def") & (summary["pod_count"] == pod_count) & (summary["qps"] == qps), stat_col
            ]
            q_val = v_q.iloc[0] if len(v_q) else None
            d_val = v_d.iloc[0] if len(v_d) else None
            ws_ch.cell(row=r, column=start_col + 1, value=float(q_val) if pd.notna(q_val) else None)
            ws_ch.cell(row=r, column=start_col + 2, value=float(d_val) if pd.notna(d_val) else None)

        header_row = start_row + 1
        last_row = start_row + 1 + len(qps_values)
        return header_row, last_row

    def add_line_chart(table_header_row: int, table_left_col: int, table_last_row: int, title: str, anchor_cell: str):
        chart = LineChart()
        chart.title = title
        chart.x_axis.title = "qps"
        chart.y_axis.title = "value (ms)"
        chart.legend.position = "r"

        cats = Reference(ws_ch, min_col=table_left_col, min_row=table_header_row + 1, max_row=table_last_row)
        data = Reference(
            ws_ch,
            min_col=table_left_col + 1,
            min_row=table_header_row,
            max_col=table_left_col + 2,
            max_row=table_last_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = False
        ws_ch.add_chart(chart, anchor_cell)

    row_cursor = 1
    left_col = 1
    for pod_count in pod_counts:
        ws_ch.cell(row=row_cursor, column=left_col, value=f"Pods {pod_count} (QoS vs Default across qps)")
        row_cursor += 1
        for metric_name, cols in metrics.items():
            for stat_label, idx in stats:
                stat_col = cols[idx]
                top = row_cursor
                header_row, last_row = write_table(top, left_col, pod_count, metric_name, stat_label, stat_col)
                anchor = ws_ch.cell(row=top, column=left_col + 4).coordinate
                add_line_chart(header_row, left_col, last_row, f"Pods {pod_count} - {metric_name} ({stat_label})", anchor)
                row_cursor = last_row + 4
        row_cursor += 2

    for ws in [ws_raw, ws_sum, ws_ch]:
        max_col = min(ws.max_column, 22)
        for col_idx in range(1, max_col + 1):
            col_letter = chr(64 + col_idx)
            max_len = 0
            for row_idx in range(1, min(ws.max_row, 200) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(out_xlsx)


def _auto_log_scale(values: np.ndarray) -> bool:
    vals = values[np.isfinite(values) & (values > 0)]
    if len(vals) < 2:
        return False
    return (np.nanmax(vals) / np.nanmin(vals)) >= 50


SYSTEM_ORDER = {"def": 0, "qos": 1, "qos_assured": 2}
SYSTEM_LABELS = {"def": "Default", "qos": "QoS", "qos_assured": "QoS-ASSURED"}
SYSTEM_COLORS = {"def": "blue", "qos": "green", "qos_assured": "orange"}


def _ordered_systems(summary: pd.DataFrame) -> list[str]:
    systems = [str(x) for x in summary["system"].dropna().unique().tolist()]
    systems = sorted(systems, key=lambda s: (SYSTEM_ORDER.get(s, 99), s))
    return systems


def _display_system(s: str) -> str:
    return SYSTEM_LABELS.get(s, s)


def _system_color(s: str) -> str:
    return SYSTEM_COLORS.get(s, "gray")


def _norm_sched_name(x: object) -> str:
    s = str(x).strip().lower()
    if "assured" in s:
        return "qos_assured"
    if "qos" in s:
        return "qos"
    if "default" in s or s == "def":
        return "def"
    return s


def build_run_count_table(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df.empty:
        return pd.DataFrame(columns=["sch", "pod_count", "qps", "n_experiments"])

    tmp = raw_df.copy()
    if "scheduler" in tmp.columns and tmp["scheduler"].notna().any():
        tmp["sch"] = tmp["scheduler"].map(_norm_sched_name)
    else:
        tmp["sch"] = tmp["system"].map(_norm_sched_name)

    counts = (
        tmp.groupby(["sch", "pod_count", "qps"], dropna=False)
        .size()
        .reset_index(name="n_experiments")
        .sort_values(["sch", "pod_count", "qps"], kind="stable")
        .reset_index(drop=True)
    )
    return counts


def plot_boxrange_charts_python_only(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
    label_avg: bool = False,
    avg_label_decimals: int = 0,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)

    metric_map = {
        "PodScheduled": ("PodScheduled_avg", "PodScheduled_p50", "PodScheduled_p99", "PodScheduled_max"),
        "ContainerReady": ("ContainerReady_avg", "ContainerReady_p50", "ContainerReady_p99", "ContainerReady_max"),
        "ServiceLatency": ("ServiceLatency_avg", "ServiceLatency_p50", "ServiceLatency_99th", "ServiceLatency_max"),
    }
    systems = [("qos", -0.18, "QoS", "green"), ("def", +0.18, "Default", "blue")]
    box_width = 0.30
    created: List[Path] = []

    for pod_count in pod_counts:
        for metric_name, (avg_col, p50_col, p99_col, max_col) in metric_map.items():
            vals = []
            for sys, _, _, _ in systems:
                sub = summary[
                    (summary["system"] == sys)
                    & (summary["pod_count"] == pod_count)
                    & (summary["qps"].isin(qps_values))
                ]
                if not sub.empty:
                    vals.extend(sub[[avg_col, p50_col, p99_col, max_col]].to_numpy().ravel().tolist())
            all_vals = np.array([v for v in vals if v is not None], dtype=float)

            use_log = _auto_log_scale(all_vals)
            if use_log and not np.all(all_vals[np.isfinite(all_vals)] > 0):
                use_log = False

            fig, ax = plt.subplots(figsize=(7.8, 4.6))
            x = np.arange(len(qps_values), dtype=float)
            ax.set_xticks(x)
            ax.set_xticklabels([str(q) for q in qps_values])
            ax.set_xlabel("qps")
            ax.set_ylabel("Time (ms)")
            ax.set_title(f"Pods {pod_count} - {metric_name} (box=p99..max, dot=avg)")

            for sys, dx, label, color in systems:
                sub = summary[
                    (summary["system"] == sys)
                    & (summary["pod_count"] == pod_count)
                    & (summary["qps"].isin(qps_values))
                ].copy()
                sub["qps"] = pd.Categorical(sub["qps"], categories=list(qps_values), ordered=True)
                sub = sub.sort_values("qps")

                proxy_done = False
                for i, qps in enumerate(qps_values):
                    row = sub[sub["qps"] == qps]
                    if row.empty:
                        continue
                    avg = row.iloc[0][avg_col]
                    p99 = row.iloc[0][p99_col]
                    mx = row.iloc[0][max_col]
                    if pd.isna(avg) and pd.isna(p99) and pd.isna(mx):
                        continue

                    xi = x[i] + dx
                    if pd.notna(p99) and pd.notna(mx) and float(mx) >= float(p99):
                        rect = Rectangle(
                            (xi - box_width / 2, float(p99)),
                            box_width,
                            float(mx) - float(p99),
                            fill=False,
                            linewidth=1.8,
                            edgecolor=color,
                            label=label if not proxy_done else None,
                        )
                        ax.add_patch(rect)
                        proxy_done = True

                    if pd.notna(avg):
                        y = float(avg)
                        ax.plot([xi], [y], marker="o", linestyle="None", color=color)
                        if label_avg:
                            fmt = f"{{:.{avg_label_decimals}f}}"
                            voff = 6 if sys == "qos" else -10
                            ax.annotate(
                                fmt.format(y),
                                (xi, y),
                                textcoords="offset points",
                                xytext=(0, voff),
                                ha="center",
                                va="bottom" if voff > 0 else "top",
                                color=color,
                                fontsize=8,
                            )

            handles, labels = ax.get_legend_handles_labels()
            if handles:
                ax.legend(loc="best")
            ax.grid(True, which="both", axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            if use_log:
                ax.set_yscale("log")

            out_path = out_dir / f"boxrange_{metric_name}_pods{pod_count}.png"
            fig.tight_layout()
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_p50_bar_xpods(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)
    systems = _ordered_systems(summary)

    metric_map = {
        "PodScheduled": "PodScheduled_p50",
        "ContainerReady": "ContainerReady_p50",
        "ServiceLatency": "ServiceLatency_p50",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, p50_col in metric_map.items():
            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), p50_col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("P50 (ms)")
            ax.set_title(f"QPS {qps} - {metric_name} (P50 bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            handles, _ = ax.get_legend_handles_labels()
            if handles:
                ax.legend(loc="best")

            out_path = out_dir / f"bar_p50_{metric_name}_qps{qps}.png"
            fig.tight_layout()
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_avg_bar_xpods(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)
    systems = _ordered_systems(summary)

    metric_map = {
        "PodScheduled": "PodScheduled_avg",
        "ContainerReady": "ContainerReady_avg",
        "ServiceLatency": "ServiceLatency_avg",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, avg_col in metric_map.items():
            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), avg_col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("AVG (ms)")
            ax.set_title(f"QPS {qps} - {metric_name} (AVG bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            handles, _ = ax.get_legend_handles_labels()
            if handles:
                ax.legend(loc="best")

            out_path = out_dir / f"bar_avg_{metric_name}_qps{qps}.png"
            fig.tight_layout()
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def plot_resource_bar_xpods(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)
    systems = _ordered_systems(summary)

    metric_map = {
        "ResourceUsage_runningPods": "ResourceUsage_runningPods",
        "ResourceUsage_cpu_total_m": "ResourceUsage_cpu_total_m",
        "ResourceUsage_mem_total_mi": "ResourceUsage_mem_total_mi",
        "ResourceUsage_net_rx_bytes_total": "ResourceUsage_net_rx_bytes_total",
        "ResourceUsage_net_tx_bytes_total": "ResourceUsage_net_tx_bytes_total",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, col in metric_map.items():
            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel(metric_name)
            ax.set_title(f"QPS {qps} - {metric_name} (bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            handles, _ = ax.get_legend_handles_labels()
            if handles:
                ax.legend(loc="best")

            out_path = out_dir / f"bar_resource_{metric_name}_qps{qps}.png"
            fig.tight_layout()
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            created.append(out_path)
    return created


def plot_containerready_counts_bar_xpods(
    raw_df: pd.DataFrame,
    out_dir: Path,
    pod_counts: Tuple[int, ...],
    qps_values: Tuple[int, ...],
    dpi: int = 200,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    summary = summarize_means(raw_df)
    systems = _ordered_systems(summary)

    metric_map = {
        "readyContainers": "ContainerReady_readyContainers",
        "expectedContainers": "ContainerReady_expectedContainers",
    }
    created: List[Path] = []

    for qps in qps_values:
        for metric_name, col in metric_map.items():
            if col not in summary.columns:
                continue

            fig, ax = plt.subplots(figsize=(8.2, 4.8))
            x = np.arange(len(pod_counts), dtype=float)
            nsys = max(len(systems), 1)
            group_w = 0.8
            width = group_w / nsys
            start = -group_w / 2 + width / 2

            for idx, sys in enumerate(systems):
                vals = []
                for pods in pod_counts:
                    v_s = summary.loc[
                        (summary["system"] == sys) & (summary["qps"] == qps) & (summary["pod_count"] == pods), col
                    ]
                    vals.append(float(v_s.iloc[0]) if len(v_s) and pd.notna(v_s.iloc[0]) else np.nan)
                ax.bar(x + start + idx * width, vals, width=width, color=_system_color(sys), label=_display_system(sys))

            ax.set_xticks(x)
            ax.set_xticklabels([str(p) for p in pod_counts])
            ax.set_xlabel("number of pods")
            ax.set_ylabel("containers")
            ax.set_title(f"QPS {qps} - ContainerReady {metric_name} (bar)")
            ax.grid(True, axis="y", linestyle="--", linewidth=0.6, alpha=0.5)
            handles, _ = ax.get_legend_handles_labels()
            if handles:
                ax.legend(loc="best")

            out_path = out_dir / f"bar_containerready_{metric_name}_qps{qps}.png"
            fig.tight_layout()
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            created.append(out_path)

    return created


def merge_metric_qps_plots(
    created_paths: List[Path],
    out_dir: Path,
    parse_regex: str,
    out_name_tpl: str,
    title_tpl: str,
    ncols: int = 2,
    dpi: int = 200,
    group_by: str = "metric",
    max_panels_per_figure: int = 4,
) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    rx = re.compile(parse_regex)
    grouped: dict[str, list[tuple[str, Path]]] = {}

    for p in created_paths:
        m = rx.match(p.name)
        if not m:
            continue
        metric = m.group("metric")
        qps = int(m.group("qps"))
        if group_by == "metric":
            grouped.setdefault(metric, []).append((qps, p))
        elif group_by == "qps":
            grouped.setdefault(str(qps), []).append((metric, p))

    merged: List[Path] = []
    for key, items in grouped.items():
        items = sorted(items, key=lambda t: t[0])
        n = len(items)
        if n == 0:
            continue

        chunk_size = max(1, max_panels_per_figure)
        pages = math.ceil(n / chunk_size)
        for page in range(pages):
            part = items[page * chunk_size : (page + 1) * chunk_size]
            pn = len(part)
            nrows = math.ceil(pn / ncols)
            fig, axes = plt.subplots(nrows, ncols, figsize=(6.4 * ncols, 4.6 * nrows))
            axes_arr = np.array(axes).reshape(-1)

            for idx, (subkey, pth) in enumerate(part):
                ax = axes_arr[idx]
                ax.imshow(plt.imread(pth))
                ax.axis("off")
                panel = chr(ord("a") + idx) if idx < 26 else str(idx + 1)
                label_text = f"({panel}) QPS {subkey}" if group_by == "metric" else f"({panel}) {subkey}"
                ax.text(
                    0.5,
                    -0.12,
                    label_text,
                    transform=ax.transAxes,
                    ha="center",
                    va="top",
                    fontsize=11,
                    clip_on=False,
                )

            for idx in range(pn, len(axes_arr)):
                axes_arr[idx].axis("off")

            title = title_tpl.format(**{group_by: key})
            if pages > 1:
                title = f"{title} (part {page + 1}/{pages})"
            fig.suptitle(title, fontsize=14, y=0.98)
            bottom = 0.15 if nrows == 1 else 0.06
            fig.subplots_adjust(left=0.04, right=0.98, bottom=bottom, top=0.90, wspace=0.04, hspace=0.35)

            base_name = out_name_tpl.format(**{group_by: key})
            if pages > 1:
                stem = Path(base_name).stem
                suffix = Path(base_name).suffix
                out_name = f"{stem}_part{page + 1}{suffix}"
            else:
                out_name = base_name
            out_path = out_dir / out_name
            fig.savefig(out_path, dpi=dpi)
            plt.close(fig)
            gc.collect()
            merged.append(out_path)

    return merged


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parse bookinfo summaries in log-result, write Excel report, and generate box-range PNG plots."
    )
    parser.add_argument(
        "--base-dir",
        default="/home/samizadeh/Downloads/swm-git/swm-benchmarking-paper-main/Experiment1/logs/",
        help="Directory containing log files and where outputs are written.",
    )
    parser.add_argument(
        "--glob",
        default="pod-summary_*.log",
        help="Glob pattern to select input logs in --base-dir.",
    )
    parser.add_argument(
        "--out-excel",
        default="pod_summary_qos_vs_def.xlsx",
        help="Output Excel filename (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--plots-dir",
        default="plots_boxrange",
        help="Directory for PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--p50-plots-dir",
        default="plots_p50_bar_xpods",
        help="Directory for P50 bar PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--avg-plots-dir",
        default="plots_avg_bar_xpods",
        help="Directory for AVG bar PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--resource-plots-dir",
        default="plots_resource_bar_xpods",
        help="Directory for ResourceUsage bar PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument(
        "--containerready-plots-dir",
        default="plots_containerready_bar_xpods",
        help="Directory for ContainerReady count bar PNG plots (inside base-dir unless absolute path).",
    )
    parser.add_argument("--dpi", type=int, default=200, help="Output DPI for generated plots.")
    parser.add_argument("--pods", default="", help="Optional comma-separated pod counts (e.g. '10,20,40').")
    parser.add_argument("--qps", default="", help="Optional comma-separated qps values (e.g. '10,100,500').")
    parser.add_argument("--label-avg", action="store_true", help="Write avg value label near each avg dot.")
    parser.add_argument("--avg-label-decimals", type=int, default=0, help="Decimals for avg labels.")
    parser.add_argument(
        "--merge-max-panels",
        type=int,
        default=4,
        help="Maximum number of single charts loaded into one merged panel image.",
    )
    args = parser.parse_args()

    base_dir = Path(args.base_dir).expanduser()
    if not base_dir.exists():
        raise SystemExit(f"ERROR: base dir does not exist: {base_dir}")

    def resolve(p: str) -> Path:
        pp = Path(p).expanduser()
        return pp if pp.is_absolute() else base_dir / pp

    log_paths = sorted(base_dir.glob(args.glob))
    if not log_paths:
        raise SystemExit(f"ERROR: no logs matched '{args.glob}' under {base_dir}")

    raw_parts = [parse_file(p) for p in log_paths]
    raw_df = pd.concat(raw_parts, ignore_index=True)
    if raw_df.empty:
        raise SystemExit("ERROR: no parseable metric blocks found in input logs.")

    out_excel = resolve(args.out_excel)
    build_excel_with_summary_and_charts(raw_df, out_excel)
    print(f"[OK] Excel written: {out_excel}")

    print("[INFO] Run-count table (tuple (sch, pod, qps) = n_experiments):")
    counts_df = build_run_count_table(raw_df)
    if counts_df.empty:
        print("no rows")
    else:
        for row in counts_df.itertuples(index=False):
            pod_display = "na" if pd.isna(row.pod_count) else int(row.pod_count)
            print(f"({row.sch}, {pod_display}, {int(row.qps)}) = {int(row.n_experiments)}")

    if args.pods.strip():
        pod_counts = tuple(int(x.strip()) for x in args.pods.split(",") if x.strip())
    else:
        pod_counts = tuple(sorted(int(x) for x in raw_df["pod_count"].dropna().unique()))

    if args.qps.strip():
        qps_values = tuple(int(x.strip()) for x in args.qps.split(",") if x.strip())
    else:
        qps_values = tuple(sorted(int(x) for x in raw_df["qps"].dropna().unique()))

    plots_dir = resolve(args.plots_dir)
    created = plot_boxrange_charts_python_only(
        raw_df=raw_df,
        out_dir=plots_dir,
        pod_counts=pod_counts,
        qps_values=qps_values,
        dpi=args.dpi,
        label_avg=args.label_avg,
        avg_label_decimals=args.avg_label_decimals,
    )
    print(f"[OK] {len(created)} plots written to: {plots_dir}")
    for p in created:
        print(f"  - {p}")

    p50_plots_dir = resolve(args.p50_plots_dir)
    created_p50 = plot_p50_bar_xpods(
        raw_df=raw_df,
        out_dir=p50_plots_dir,
        pod_counts=pod_counts,
        qps_values=qps_values,
        dpi=args.dpi,
    )
    print(f"[OK] {len(created_p50)} P50 bar plots written to: {p50_plots_dir}")
    merged_p50 = merge_metric_qps_plots(
        created_paths=created_p50,
        out_dir=p50_plots_dir,
        parse_regex=r"^bar_p50_(?P<metric>.+)_qps(?P<qps>\d+)\.png$",
        out_name_tpl="panel_2x2_bar_p50_{metric}.png",
        title_tpl="{metric} - P50 bars across QPS",
        dpi=args.dpi,
        max_panels_per_figure=args.merge_max_panels,
    )
    print(f"[OK] {len(merged_p50)} merged 2x2 P50 panels written to: {p50_plots_dir}")

    avg_plots_dir = resolve(args.avg_plots_dir)
    created_avg = plot_avg_bar_xpods(
        raw_df=raw_df,
        out_dir=avg_plots_dir,
        pod_counts=pod_counts,
        qps_values=qps_values,
        dpi=args.dpi,
    )
    print(f"[OK] {len(created_avg)} AVG bar plots written to: {avg_plots_dir}")
    merged_avg = merge_metric_qps_plots(
        created_paths=created_avg,
        out_dir=avg_plots_dir,
        parse_regex=r"^bar_avg_(?P<metric>.+)_qps(?P<qps>\d+)\.png$",
        out_name_tpl="panel_2x2_bar_avg_{metric}.png",
        title_tpl="{metric} - AVG bars across QPS",
        dpi=args.dpi,
        max_panels_per_figure=args.merge_max_panels,
    )
    print(f"[OK] {len(merged_avg)} merged 2x2 AVG panels written to: {avg_plots_dir}")

    resource_plots_dir = resolve(args.resource_plots_dir)
    created_resource = plot_resource_bar_xpods(
        raw_df=raw_df,
        out_dir=resource_plots_dir,
        pod_counts=pod_counts,
        qps_values=qps_values,
        dpi=args.dpi,
    )
    print(f"[OK] {len(created_resource)} ResourceUsage bar plots written to: {resource_plots_dir}")
    merged_resource = merge_metric_qps_plots(
        created_paths=created_resource,
        out_dir=resource_plots_dir,
        parse_regex=r"bar_resource_(?P<metric>.*?)_qps(?P<qps>\d+)\.png",
        out_name_tpl="panel_2x2_bar_resource_{metric}.png",
        title_tpl="Resource Usage: {metric}",
        ncols=2,
        dpi=args.dpi,
        max_panels_per_figure=args.merge_max_panels,
    )
    print(f"[OK] {len(merged_resource)} merged 2x2 resource panels written to: {resource_plots_dir}")

    containerready_plots_dir = resolve(args.containerready_plots_dir)
    created_cr = plot_containerready_counts_bar_xpods(
        raw_df=raw_df,
        out_dir=containerready_plots_dir,
        pod_counts=pod_counts,
        qps_values=qps_values,
        dpi=args.dpi,
    )
    print(f"[OK] {len(created_cr)} ContainerReady count bar plots written to: {containerready_plots_dir}")
    merged_cr = merge_metric_qps_plots(
        created_paths=created_cr,
        out_dir=containerready_plots_dir,
        parse_regex=r"bar_containerready_(?P<metric>.*?)_qps(?P<qps>\d+)\.png",
        out_name_tpl="panel_2x2_bar_containerready_{metric}.png",
        title_tpl="ContainerReady counts: {metric}",
        ncols=2,
        dpi=args.dpi,
        max_panels_per_figure=args.merge_max_panels,
    )
    print(f"[OK] {len(merged_cr)} merged 2x2 ContainerReady count panels written to: {containerready_plots_dir}")


if __name__ == "__main__":
    main()
